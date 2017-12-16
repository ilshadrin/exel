'''
читаем из файла data_1 данные  и записываем их в эксель

'''

import csv
from openpyxl import  Workbook  

def read_csv(filename):
    data=[]
    with open(filename, 'r', encoding='utf-8') as f:
        fields=['Stantion', 'street']
        reader = csv.DictReader(f, fields, delimiter=',')
        for row in reader:
            data.append(row)
    return data



def excel_write(data):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "TestTestTest"    # Задаем заголовок

    worksheet.cell(row=1, column=1).value='Станция'  #заполняем лист значениями. cell - ячейка, row - строка,  column - колонка. value - знаение,  =Станция - значение 1ой яейки
    worksheet.cell(row=1, column=2).value='Улица'    # Улица - значение второй ячейки

    row=2
    for item in data:                                 # записываем данные из файла data_1 в эксель с помощью цикла 
        worksheet.cell(row=row, column=1).value=item['Stantion'] #ключи словаря fields
        worksheet.cell(row=row, column=2).value=item['street']
        row+= 1

    workbook.save('ExelTest.xlsx')   # создаем эксель файл

csv_data=read_csv('data_1.csv')
excel_write(csv_data)

